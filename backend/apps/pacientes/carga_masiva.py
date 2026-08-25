from datetime import datetime

import openpyxl

from apps.pacientes.serializers import PacienteSerializer

HOJA_ESPERADA = "Pacientes"
FORMATOS_FECHA = ("%d/%m/%Y", "%d-%m-%Y")


class CargaMasivaError(Exception):
    """Error que impide procesar el archivo por completo (no un error de fila)."""


def _fila_vacia(row) -> bool:
    return row is None or all(c is None or str(c).strip() == "" for c in row)


def _parse_fecha(valor):
    """Devuelve (fecha, error). openpyxl ya entrega datetime si la celda tiene formato de fecha."""
    if valor is None or str(valor).strip() == "":
        return None, "la fecha de nacimiento es obligatoria"
    if isinstance(valor, datetime):
        return valor.date(), None
    texto = str(valor).strip()
    for fmt in FORMATOS_FECHA:
        try:
            return datetime.strptime(texto, fmt).date(), None
        except ValueError:
            continue
    return None, f"fecha '{texto}' invalida, usa el formato DD/MM/AAAA"


def procesar_carga_masiva_pacientes(archivo, *, clinica, context: dict) -> dict:
    """
    Lee el archivo de carga masiva de pacientes (hoja "Pacientes", columnas
    tipo_documento/numero_documento/nombres/apellidos/fecha_nacimiento/sexo/
    telefono/email) y crea un Paciente por fila valida, reutilizando
    PacienteSerializer para las mismas validaciones que la creacion manual
    (duplicados, formato de documento, rango de fecha, etc).

    No falla en bloque: las filas invalidas se reportan y las validas se
    crean igual, para no perder una carga completa por unos pocos errores.
    """
    try:
        wb = openpyxl.load_workbook(archivo, data_only=True)
    except Exception as exc:
        raise CargaMasivaError(f"No se pudo leer el archivo: {exc}") from exc

    if HOJA_ESPERADA not in wb.sheetnames:
        raise CargaMasivaError(f"El archivo debe tener una hoja llamada '{HOJA_ESPERADA}'.")
    ws = wb[HOJA_ESPERADA]

    creados = 0
    errores: list[dict] = []
    documentos_en_archivo: set[tuple[str, str]] = set()

    for fila_idx, row in enumerate(ws.iter_rows(min_row=2, max_col=8, values_only=True), start=2):
        if _fila_vacia(row):
            continue

        valores = (list(row) + [None] * 8)[:8]
        (
            tipo_documento_raw, numero_documento_raw, nombres_raw, apellidos_raw,
            fecha_raw, sexo_raw, telefono_raw, email_raw,
        ) = valores

        tipo_documento = str(tipo_documento_raw).strip().upper() if tipo_documento_raw else ""
        numero_documento = str(numero_documento_raw).strip() if numero_documento_raw else ""
        nombres = str(nombres_raw).strip() if nombres_raw else ""
        apellidos = str(apellidos_raw).strip() if apellidos_raw else ""
        sexo = str(sexo_raw).strip().upper() if sexo_raw else ""
        telefono = str(telefono_raw).strip() if telefono_raw else ""
        email = str(email_raw).strip() if email_raw else ""

        fila_errores = []
        fecha_nacimiento, fecha_error = _parse_fecha(fecha_raw)
        if fecha_error:
            fila_errores.append(fecha_error)

        clave = (tipo_documento, numero_documento)
        if numero_documento and clave in documentos_en_archivo:
            fila_errores.append("documento duplicado dentro del archivo")
        else:
            documentos_en_archivo.add(clave)

        if fila_errores:
            errores.append({
                "fila": fila_idx,
                "documento": numero_documento or None,
                "mensaje": "; ".join(fila_errores),
            })
            continue

        data = {
            "clinica": str(clinica.id),
            "tipo_documento": tipo_documento,
            "numero_documento": numero_documento,
            "nombres": nombres,
            "apellidos": apellidos,
            "fecha_nacimiento": fecha_nacimiento.isoformat(),
            "sexo": sexo,
            "telefono": telefono,
            "email": email,
            # Carga masiva = migracion de pacientes existentes de otro sistema;
            # la clinica atestigua que ya cuenta con esa autorizacion previa.
            "autoriza_datos": True,
        }
        serializer = PacienteSerializer(data=data, context=context)
        if serializer.is_valid():
            serializer.save()
            creados += 1
        else:
            mensaje = " | ".join(
                f"{campo}: {'; '.join(str(e) for e in errs)}"
                for campo, errs in serializer.errors.items()
            )
            errores.append({"fila": fila_idx, "documento": numero_documento or None, "mensaje": mensaje})

    return {
        "total_filas": creados + len(errores),
        "creados": creados,
        "errores": errores,
    }
